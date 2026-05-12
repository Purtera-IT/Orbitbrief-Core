"""Regenerate ``brains/data/briefing_configs.yaml`` from the intake workbook.

Source: ``AWESOME_CHASE_orbitbrief_domain_only_schema_intake_workbook_v4.xlsx``.

For each briefing-shaped domain (wireless, low_voltage_cabling,
rack_and_stack, datacenter, imac) we extract:

* ``operating_rules`` — boolean flags from the workbook (e.g.
  ``do_not_invent_facts``, ``use_assumptions_for_supported_gaps_only``).
* ``normalization`` — controlled vocabularies + abbreviation maps
  (e.g. wireless's ``survey_type_labels`` and ``common_wireless_terms``).
* ``fields`` — the canonical 9-field POST schema with per-field
  guidance bullets unioned across all source-type variants.

For domains where the workbook has empty per-field guidance
(currently low_voltage_cabling, rack_and_stack, datacenter, imac),
we fall back to ``DEFAULT_GUIDANCE`` plus subdomain notes mined
from the ``01_INDEX`` sheet so brains have something concrete to
prompt with. Refresh this file when the workbook is filled in.
"""
from __future__ import annotations

import json
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl
import yaml


DOMAINS: dict[str, str] = {
    "wireless": "D03_wireless",
    "low_voltage_cabling": "D05_low_voltage_cabling",
    "rack_and_stack": "D06_rack_and_stack",
    "datacenter": "D07_datacenter",
    "imac": "D08_imac",
}

CANONICAL_FIELDS: tuple[str, ...] = (
    "scope_overview",
    "detailed_scope_of_services",
    "deliverables",
    "assumptions",
    "customer_responsibilities",
    "out_of_scope",
    "risks_or_dependencies",
    "completion_criteria",
    "open_items",
)


# Default guidance reused across domains where the workbook has empty
# per-field guidance. Reviewers should still feel the per-domain flavor,
# so the per-domain defaults below LAYER on top of these.
DEFAULT_GUIDANCE: dict[str, list[str]] = {
    "scope_overview": [
        "1–3 sentences describing the engagement at a glance",
        "Reference the customer site or facility if known",
        "State delivery model (onsite, remote, hybrid) if known",
    ],
    "detailed_scope_of_services": [
        "Concrete activities the engagement performs",
        "One bullet per executable activity, no nested prose",
        "Use execution-ready statements that a PM can sequence",
    ],
    "deliverables": [
        "Customer-facing tangible outputs only",
        "Do not mix services with deliverables",
    ],
    "assumptions": [
        "Atomic, testable assumptions only",
        "Do not invent scope; mark gaps as open_items instead",
    ],
    "customer_responsibilities": [
        "Required customer actions, inputs, permissions, files",
        "Be explicit about timing where it affects sequencing",
    ],
    "out_of_scope": [
        "Explicit exclusions clarifying scope boundary",
        "Avoid exclusions unrelated to the opportunity",
    ],
    "risks_or_dependencies": [
        "Execution risks, dependencies, limiting conditions",
        "Include unknowns that may affect schedule, quality, or pricing",
    ],
    "completion_criteria": [
        "Objective indicators that the engagement is complete",
        "Align directly to the listed services + deliverables",
    ],
    "open_items": [
        "Specific unresolved items that block quote finalization",
        "Use for missing, conflicting, or unclear intake details",
    ],
}


# Per-domain hand-curated extras. Sourced from the 01_INDEX sheet
# (subdomain notes) and operator knowledge. Refresh when the workbook
# fills in domain-specific guidance.
DOMAIN_DEFAULT_GUIDANCE: dict[str, dict[str, list[str]]] = {
    "low_voltage_cabling": {
        "detailed_scope_of_services": [
            "Differentiate copper drops, fiber runs, DMARC extension, IDF homerun",
            "Call out testing / certification standards (BICSI, TIA-568)",
        ],
        "assumptions": [
            "What is assumed existing vs new (drops, pathways, patch panels)",
            "Pathway availability + ceiling height assumptions",
        ],
        "out_of_scope": [
            "Common: existing copper reuse not certified by us",
            "Common: pathway construction (J-hooks vs conduit)",
        ],
        "open_items": [
            "Missing: total drop count by room",
            "Missing: as-built reference if reusing existing",
        ],
    },
    "rack_and_stack": {
        "detailed_scope_of_services": [
            "Device counts/types, rack elevations, in-rack patch matrix",
            "Cable management, dressing, labeling standards",
        ],
        "deliverables": [
            "Updated rack elevations, patch matrix, labeling export",
            "Photo set per cabinet if requested",
        ],
        "customer_responsibilities": [
            "Provide rack elevations, patch matrix, asset tags",
            "Confirm power whip availability per cabinet",
        ],
        "open_items": [
            "Missing: per-cabinet power whip type",
            "Missing: device weight per RU for floor loading",
        ],
    },
    "datacenter": {
        "detailed_scope_of_services": [
            "Differentiate rack/stack vs power/patching vs decommission",
            "Power requirements (208V vs 120V), PDU/CDU class, RU constraints",
        ],
        "assumptions": [
            "Power capacity available at each cabinet",
            "Existing patching matrix is current",
        ],
        "open_items": [
            "Missing: facility access windows + escort policy",
            "Missing: RU constraints per cabinet",
        ],
    },
    "imac": {
        "detailed_scope_of_services": [
            "Refresh / swap / move / add / change counts per device class",
            "User-device mapping, image template, migration dependencies",
        ],
        "customer_responsibilities": [
            "User schedule + cutover availability",
            "Old asset disposition (return-to-vendor / ITAD)",
        ],
        "open_items": [
            "Missing: device-to-user map",
            "Missing: legacy asset disposition path",
        ],
    },
    "wireless": {
        # Wireless workbook is rich; defaults stay empty so workbook
        # guidance dominates.
    },
}


def _index_notes(wb) -> dict[str, list[str]]:
    """Pull subdomain notes per domain id from the ``01_INDEX`` sheet."""
    out: dict[str, list[str]] = defaultdict(list)
    ws = wb["01_INDEX"]
    for r, row in enumerate(ws.iter_rows(values_only=True)):
        if r < 5:
            continue
        domain, subdomain, did, _, alias, notes = (row + (None,) * 6)[:6]
        if not did:
            continue
        if notes and str(notes).strip():
            out[str(did).strip()].append(str(notes).strip())
    return dict(out)


def _extract_post_schemas(sheet) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    current_section = ""
    for r, row in enumerate(sheet.iter_rows(values_only=True)):
        first = row[0] if row else None
        if first and isinstance(first, str) and first.strip() and not first.startswith(
            "File type"
        ):
            if "—" in first or first.startswith(
                ("transcript_", "site_roster_", "rfp_", "vendor_quote_")
            ):
                current_section = first.strip()
                continue
        if len(row) >= 5 and row[4] and isinstance(row[4], str) and row[4].strip().startswith("{"):
            try:
                doc = json.loads(row[4])
                doc["_artifact_label"] = (row[0] or "").strip() if row[0] else ""
                doc["_section"] = current_section
                out.append(doc)
            except json.JSONDecodeError:
                pass
    return out


def _build_field_guidance(
    schemas: list[dict[str, Any]], domain_id: str, index_notes: list[str]
) -> dict[str, list[str]]:
    """Union of guidance bullets across all source-type variants per field."""
    per_field: dict[str, set[str]] = defaultdict(set)

    for s in schemas:
        for root_key in ("output", "fields", "fixed_fields"):
            block = s.get(root_key)
            if not isinstance(block, dict):
                continue
            for fname, fval in block.items():
                if isinstance(fval, dict):
                    for g in (fval.get("guidance") or []):
                        if isinstance(g, str) and g.strip():
                            per_field[fname].add(g.strip())

    # Layer in per-domain defaults.
    for fname, bullets in DOMAIN_DEFAULT_GUIDANCE.get(domain_id, {}).items():
        for b in bullets:
            per_field.setdefault(fname, set()).add(b)
    # Layer in cross-domain defaults.
    for fname, bullets in DEFAULT_GUIDANCE.items():
        for b in bullets:
            per_field.setdefault(fname, set()).add(b)
    # Drop any field that isn't canonical (paranoia).
    canonical = set(CANONICAL_FIELDS)
    for fname in list(per_field.keys()):
        if fname not in canonical:
            per_field.pop(fname)
    return {f: sorted(per_field[f]) for f in CANONICAL_FIELDS if f in per_field}


def build_bundle(workbook_path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(str(workbook_path), data_only=True)
    index_notes_by_domain = _index_notes(wb)
    bundle: dict[str, Any] = {
        "_doc": (
            "OrbitBrief domain briefing configs extracted from "
            "AWESOME_CHASE intake workbook v4. Workbook fills override "
            "DEFAULT_GUIDANCE; per-domain defaults layer in between."
        ),
        "version": "v4",
        "canonical_fields": list(CANONICAL_FIELDS),
        "domains": {},
    }

    for domain_id, sheet_name in DOMAINS.items():
        if sheet_name not in wb.sheetnames:
            continue
        schemas = _extract_post_schemas(wb[sheet_name])
        # Operating rules: union; first occurrence wins.
        operating_rules: dict[str, Any] = {}
        for s in schemas:
            for k, v in (s.get("operating_rules") or {}).items():
                operating_rules.setdefault(k, v)
        # Normalization (vocabularies / abbreviation maps).
        normalization: dict[str, Any] = {}
        for s in schemas:
            for k, v in (s.get("normalization") or {}).items():
                if isinstance(v, list):
                    existing = normalization.setdefault(k, [])
                    for item in v:
                        if item not in existing:
                            existing.append(item)
                elif isinstance(v, dict):
                    target = normalization.setdefault(k, {})
                    for nk, nv in v.items():
                        target.setdefault(nk, nv)

        artifact_labels = sorted(
            {(s.get("_artifact_label") or "").strip() for s in schemas if s.get("_artifact_label")}
        )

        bundle["domains"][domain_id] = {
            "display_name": domain_id.replace("_", " ").title(),
            "operating_rules": operating_rules,
            "normalization": normalization,
            "fields": _build_field_guidance(
                schemas, domain_id, index_notes_by_domain.get(domain_id, [])
            ),
            "artifact_labels": artifact_labels,
            "schemas_extracted": len(schemas),
            "subdomain_notes": index_notes_by_domain.get(domain_id, []),
        }

    return bundle


def main(argv: list[str]) -> int:
    if len(argv) != 2:
        print(
            "usage: python tools/extract_briefing_configs.py <workbook.xlsx>",
            file=sys.stderr,
        )
        return 2
    src = Path(argv[1])
    if not src.is_file():
        print(f"workbook not found: {src}", file=sys.stderr)
        return 1
    out = (
        Path(__file__).resolve().parent.parent
        / "src"
        / "orbitbrief_core"
        / "brains"
        / "data"
        / "briefing_configs.yaml"
    )
    out.parent.mkdir(parents=True, exist_ok=True)
    bundle = build_bundle(src)
    out.write_text(
        yaml.safe_dump(bundle, sort_keys=False, allow_unicode=True, width=120),
        encoding="utf-8",
    )
    print(f"wrote {out} ({out.stat().st_size:,} bytes)")
    for d, cfg in bundle["domains"].items():
        guidance_count = sum(len(v) for v in cfg["fields"].values())
        print(
            f"  {d}: {cfg['schemas_extracted']} schemas → "
            f"{len(cfg['fields'])} fields, {guidance_count} guidance bullets, "
            f"{len(cfg['normalization'])} normalization groups"
        )
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main(sys.argv))
