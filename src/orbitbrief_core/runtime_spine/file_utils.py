from __future__ import annotations

"""Legacy utility parsers.

New parser-first modules are under `runtime_spine/parsers/`.
Keep this file only as backward compatibility until full cutover.
"""

import csv
import hashlib
import io
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


def sha256_file(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def sniff_modality(path: Path) -> tuple[str, list[str]]:
    ext = path.suffix.lower()
    lowered = path.name.lower()
    reasons = [f"extension={ext or 'none'}"]
    if "email" in lowered and ext in {".txt", ".md", ".docx"}:
        return "email_export", reasons + ["filename suggests email export"]
    if "pasted" in lowered or "blob" in lowered:
        return "pasted_notes", reasons + ["filename suggests pasted notes/text blob"]
    if ext == ".txt":
        return "txt", reasons
    if ext == ".md":
        return "md", reasons
    if ext == ".docx":
        return "docx", reasons
    if ext == ".xlsx":
        return "xlsx", reasons
    if ext == ".xls":
        return "xls", reasons
    if ext == ".csv":
        return "csv", reasons
    if ext == ".pdf":
        if "dwg" in lowered:
            return "dwg_export_pdf", reasons + ["filename contains dwg"]
        if "image" in lowered or "scan" in lowered:
            return "image_pdf", reasons + ["filename suggests image-based pdf"]
        return "pdf", reasons
    return ext.lstrip(".") or "unknown", reasons


def read_textual_file(path: Path, modality: str) -> str:
    if modality in {"txt", "md", "pasted_notes", "text_blob", "email_export"}:
        return path.read_text(encoding="utf-8", errors="ignore")
    if modality == "docx":
        return extract_docx_text(path)
    raise ValueError(f"Unsupported textual modality: {modality}")


def extract_docx_text(path: Path) -> str:
    with zipfile.ZipFile(path, "r") as zf:
        xml = zf.read("word/document.xml")
    root = ET.fromstring(xml)
    texts = []
    for node in root.iter():
        if node.tag.endswith("}t") and node.text:
            texts.append(node.text)
        if node.tag.endswith("}p"):
            texts.append("\n")
    return "".join(texts)


def extract_pdf_text(path: Path) -> str:
    raw = path.read_bytes()
    text_chunks = []
    for match in re.findall(rb"\(([^()]*)\)\s*Tj", raw):
        try:
            text_chunks.append(match.decode("latin1"))
        except Exception:
            continue
    if text_chunks:
        return "\n".join(text_chunks)
    try:
        return raw.decode("latin1", errors="ignore")
    except Exception:
        return ""


def pdf_page_count(path: Path) -> int:
    raw = path.read_bytes()
    count = len(re.findall(rb"/Type\s*/Page\b", raw))
    return max(1, count)


def load_csv_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8", errors="ignore", newline="") as fh:
        reader = csv.DictReader(fh)
        headers = reader.fieldnames or []
        rows = [dict(row) for row in reader]
    return headers, rows


def load_xlsx_rows(path: Path) -> tuple[str, list[str], list[dict[str, str]]]:
    wb = load_workbook(path, data_only=True)
    best_sheet = None
    best_rows = -1
    for ws in wb.worksheets:
        values = list(ws.iter_rows(values_only=True))
        non_empty = [row for row in values if any(cell not in (None, "") for cell in row)]
        if len(non_empty) > best_rows:
            best_sheet = ws
            best_rows = len(non_empty)
    if best_sheet is None:
        return "Sheet1", [], []
    values = list(best_sheet.iter_rows(values_only=True))
    non_empty = [row for row in values if any(cell not in (None, "") for cell in row)]
    if not non_empty:
        return best_sheet.title, [], []
    headers = [str(cell or "").strip() for cell in non_empty[0]]
    rows = []
    for raw_row in non_empty[1:]:
        row = {}
        for idx, header in enumerate(headers):
            key = header or f"column_{idx+1}"
            row[key] = "" if idx >= len(raw_row) or raw_row[idx] is None else str(raw_row[idx])
        rows.append(row)
    return best_sheet.title, headers, rows


def simple_header_normalize(header: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "_", header.lower()).strip("_")
    return normalized or "unnamed_column"


def split_paragraphs(text: str) -> list[str]:
    paras = [chunk.strip() for chunk in re.split(r"\n\s*\n", text.replace("\r", "\n"))]
    return [p for p in paras if p]


def text_lines(text: str) -> list[str]:
    return [line.strip() for line in text.replace("\r", "\n").split("\n") if line.strip()]


def synthetic_minimal_pdf(text: str) -> bytes:
    body = f"""%PDF-1.4
1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj
2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj
3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 300 300] /Contents 4 0 R >> endobj
4 0 obj << /Length 44 >> stream
BT
/F1 12 Tf
72 720 Td
({text}) Tj
ET
endstream
endobj
trailer << /Root 1 0 R >>
%%EOF
"""
    return body.encode("latin1", errors="ignore")
