from __future__ import annotations

import csv
import hashlib
import re
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from .config import post_schema_ref
from .contracts import EvidenceObject, FieldClaim, PageRefOrSheetRef, ReviewFlag, RoleGraph, SourceRef
from .mapping import resolve_alias
from .mapping_models import HeaderBundle, HeaderPosition, ValueProfile
from .parsers.professional_services.text_narrative import TextNarrativeParser


def _now() -> str:
    return "2026-04-06T00:00:00Z"


def _artifact_hash(path: Path) -> str:
    data = path.read_bytes() if path.exists() else b""
    return hashlib.sha256(data).hexdigest()


def _source_ref(path: Path) -> SourceRef:
    return SourceRef(
        artifact_id=path.stem,
        artifact_name=path.name,
        artifact_path=str(path),
        artifact_hash=_artifact_hash(path),
    )


def _role_graph(path: Path, role_id: str, modality: str, summary: str) -> RoleGraph:
    return RoleGraph(
        id=f"graph_{role_id}_{path.stem}",
        domain_id="professional_services",
        role_id=role_id,
        modality=modality,
        source_artifact_id=path.stem,
        source_ref=_source_ref(path),
        summary=summary[:400],
        confidence=0.9,
        created_at=_now(),
    )


def _review_flag(role_id: str, modality: str, code: str, message: str, *, requires_32b: bool = False) -> ReviewFlag:
    return ReviewFlag(
        id=f"flag_{role_id}_{code}",
        domain_id="professional_services",
        role_id=role_id,
        modality=modality,
        severity="medium",
        code=code,
        message=message,
        created_at=_now(),
        requires_32b=requires_32b,
    )


def _field_claim(role_id: str, modality: str, target_layer: str, field_name: str, field_path: str, value, schema_ref: str, evidence_refs: list[str]) -> FieldClaim:
    return FieldClaim(
        id=f"claim_{role_id}_{field_name}_{len(evidence_refs)}_{abs(hash(str(value))) % 100000}",
        domain_id="professional_services",
        role_id=role_id,
        modality=modality,
        target_layer=target_layer,
        field_name=field_name,
        field_path=field_path,
        candidate_value=value,
        normalized_value=value,
        schema_ref=schema_ref,
        evidence_refs=evidence_refs,
        confidence=0.9,
        claim_status="asserted",
        created_at=_now(),
    )


def _read_docx_text(path: Path) -> str:
    with ZipFile(path) as zf:
        xml = zf.read("word/document.xml").decode("utf-8", errors="ignore")
    xml = re.sub(r"</w:p>", "\n", xml)
    xml = re.sub(r"<[^>]+>", "", xml)
    return xml


def _read_pdf_text(path: Path) -> str:
    raw = path.read_bytes().decode("latin-1", errors="ignore")
    parts = re.findall(r"\((.*?)\)", raw, flags=re.S)
    if parts:
        cleaned = []
        for part in parts:
            text = part.replace("\\(", "(").replace("\\)", ")").replace("\\\\", "\\")
            cleaned.append(text)
        return "\n".join(cleaned)
    return raw


def _read_text_for_modality(path: Path, modality: str) -> str:
    modality = modality.lower()
    if modality == "docx":
        return _read_docx_text(path)
    if modality == "pdf":
        return _read_pdf_text(path)
    return path.read_text(encoding="utf-8", errors="ignore")


def _extract_transcript_claims(text: str, modality: str) -> tuple[list[FieldClaim], list[EvidenceObject]]:
    blocks = TextNarrativeParser().parse(Path("/dev/null"), "txt") if False else None
    del blocks
    evidence: list[EvidenceObject] = []
    claims: list[FieldClaim] = []
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    for idx, line in enumerate(lines, start=1):
        evidence_ref = f"seg_{idx:04d}"
        evidence.append(
            EvidenceObject(
                object_id=evidence_ref,
                object_type="NarrativeBlock",
                text=line,
                page_ref_or_sheet_ref=PageRefOrSheetRef(name="body"),
                metadata={"ordinal": idx},
            )
        )
        lower = line.lower()
        if lower.startswith("project summary:"):
            value = line.split(":", 1)[1].strip()
            claims.append(_field_claim("transcript_or_notes", modality, "pre_field", "project_summary", "project_summary", value, f"transcript_or_notes.{modality}.pre", [evidence_ref]))
            if modality == "docx":
                claims.append(_field_claim("transcript_or_notes", modality, "post_hint", "scope_overview", "scope_overview", value, "transcript_or_notes.docx.post.alias", [evidence_ref]))
        elif lower.startswith("assumption:"):
            value = line.split(":", 1)[1].strip()
            claims.append(_field_claim("transcript_or_notes", modality, "pre_field", "known_assumptions", "known_assumptions[]", [value], f"transcript_or_notes.{modality}.pre", [evidence_ref]))
        elif lower.startswith("exclusion:") or lower.startswith("out of scope:"):
            value = line.split(":", 1)[1].strip()
            claims.append(_field_claim("transcript_or_notes", modality, "pre_field", "known_exclusions", "known_exclusions[]", [value], f"transcript_or_notes.{modality}.pre", [evidence_ref]))
        elif lower.startswith("open question:"):
            value = line.split(":", 1)[1].strip()
            claims.append(_field_claim("transcript_or_notes", modality, "pre_field", "open_questions", "open_questions[]", [value], f"transcript_or_notes.{modality}.pre", [evidence_ref]))
            if modality == "docx":
                claims.append(_field_claim("transcript_or_notes", modality, "post_hint", "open_items", "open_items[]", [value], "transcript_or_notes.docx.post.alias", [evidence_ref]))
        elif re.search(r"\binstall\b", lower):
            claims.append(_field_claim("transcript_or_notes", modality, "pre_field", "scope_tasks_requested", "scope_tasks_requested[]", [line], f"transcript_or_notes.{modality}.pre", [evidence_ref]))
        elif re.search(r"\b\d+\s+sites?\b", lower):
            match = re.search(r"(\d+)\s+sites?", lower)
            if match:
                claims.append(_field_claim("transcript_or_notes", modality, "pre_field", "site_count", "site_count", int(match.group(1)), f"transcript_or_notes.{modality}.pre", [evidence_ref]))
    return claims, evidence


def ingest_transcript_or_notes(path: Path, modality: str) -> dict:
    modality_l = modality.lower()
    text = _read_text_for_modality(path, modality_l if modality_l != "email_export" else "txt")
    claims, evidence = _extract_transcript_claims(text, modality_l)
    return {
        "role_graph": _role_graph(path, "transcript_or_notes", modality_l, text),
        "evidence_objects": evidence,
        "field_claims": claims,
        "review_flags": [],
    }


def _pdf_summary_text(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def ingest_drawing_packet(path: Path, modality: str) -> dict:
    text = _read_pdf_text(path)
    summary = _pdf_summary_text(text)
    evidence = [
        EvidenceObject(
            object_id="sheet_0001",
            object_type="SheetObject",
            text=summary,
            page_ref_or_sheet_ref=PageRefOrSheetRef(name=path.stem, page_number=1),
            metadata={"title": path.stem},
        ),
        EvidenceObject(
            object_id="crop_0001",
            object_type="ImageCrop",
            text=None,
            page_ref_or_sheet_ref=PageRefOrSheetRef(name=path.stem, page_number=1),
            metadata={"region": "full_page"},
        ),
    ]
    claims: list[FieldClaim] = []
    refs = [obj.object_id for obj in evidence]
    if re.search(r"\bsite\b", text, flags=re.I):
        match = re.search(r"site\s+([^\n]+)", text, flags=re.I)
        value = match.group(1).strip() if match else summary
        claims.append(_field_claim("drawing_packet", modality, "pre_field", "location_details", "location_details", value, "drawing_packet.pdf.pre", refs))
    if re.search(r"access|badge", text, flags=re.I):
        claims.append(_field_claim("drawing_packet", modality, "pre_field", "access_constraints", "access_constraints[]", [summary], "drawing_packet.pdf.pre", refs))
    qty_match = re.search(r"(\d+)\s+(racks?|aps?|switches?|cabinets?)", text, flags=re.I)
    if qty_match:
        claims.append(_field_claim("drawing_packet", modality, "pre_field", "known_quantities", "known_quantities[]", [{"quantity": int(qty_match.group(1)), "unit": qty_match.group(2)}], "drawing_packet.pdf.pre", refs))
    if re.search(r"open question|\?", text, flags=re.I):
        claims.append(_field_claim("drawing_packet", modality, "pre_field", "open_questions", "open_questions[]", [summary], "drawing_packet.pdf.pre", refs))
    if re.search(r"deliverable|as built|as-built", text, flags=re.I):
        claims.append(_field_claim("drawing_packet", modality, "pre_field", "deliverables_needed", "deliverables_needed[]", [summary], "drawing_packet.pdf.pre", refs))
    if re.search(r"test requirement|testing", text, flags=re.I):
        claims.append(_field_claim("drawing_packet", modality, "pre_field", "testing_requirements", "testing_requirements[]", [summary], "drawing_packet.pdf.pre", refs))
    review_flags = [_review_flag("drawing_packet", modality, "requires_32b_path", "Drawing packets require 32b review path", requires_32b=True)]
    return {
        "role_graph": _role_graph(path, "drawing_packet", modality, summary),
        "evidence_objects": evidence,
        "field_claims": claims,
        "review_flags": review_flags,
    }


def _build_header_bundle(header: str, sample_values: list[str], *, modality: str, column_index: int) -> HeaderBundle:
    looks_like_count = bool(sample_values) and all(str(v).isdigit() for v in sample_values if str(v).strip())
    looks_like_date = bool(sample_values) and all("-" in str(v) for v in sample_values if str(v).strip())
    dominant = "date" if looks_like_date else ("count" if looks_like_count else "text")
    return HeaderBundle(
        role_id="site_roster_spreadsheet",
        domain_id="professional_services",
        modality=modality,
        header_raw=header,
        header_normalized=header.lower().replace("/", " ").replace("#", " ").strip(),
        sheet_name="Sites",
        neighbor_headers=[],
        sample_values=[str(v) for v in sample_values if v is not None],
        value_profile=ValueProfile(
            dominant_type=dominant,
            distinct_ratio=1.0,
            null_ratio=0.0,
            looks_like_date=looks_like_date,
            looks_like_count=looks_like_count,
        ),
        header_position=HeaderPosition(sheet_index=1, column_index=column_index),
    )


def ingest_site_roster_spreadsheet(path: Path, modality: str) -> dict:
    modality_l = modality.lower()
    if modality_l == "xls":
        return {
            "role_graph": _role_graph(path, "site_roster_spreadsheet", modality_l, "legacy xls roster"),
            "evidence_objects": [],
            "field_claims": [],
            "mapping_decisions": [],
            "review_flags": [_review_flag("site_roster_spreadsheet", modality_l, "xls_not_yet_supported", "Legacy XLS is routed to review")],
        }

    if modality_l == "csv":
        with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
            reader = csv.reader(handle)
            rows = list(reader)
        headers = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []
        sheet_name = "CSV"
    else:
        wb = load_workbook(path)
        ws = wb.active
        sheet_name = ws.title
        values = list(ws.iter_rows(values_only=True))
        headers = [str(v) if v is not None else "" for v in (values[0] if values else [])]
        data_rows = [["" if v is None else str(v) for v in row] for row in values[1:]]

    table_ref = PageRefOrSheetRef(name=sheet_name)
    evidence_objects = [
        EvidenceObject(
            object_id="table_0001",
            object_type="TableObject",
            text=None,
            page_ref_or_sheet_ref=table_ref,
            metadata={"headers": headers},
        )
    ]
    row_claim_value = []
    for idx, row in enumerate(data_rows, start=1):
        row_dict = {headers[col]: row[col] if col < len(row) else "" for col in range(len(headers))}
        row_claim_value.append(row_dict)
        evidence_objects.append(
            EvidenceObject(
                object_id=f"row_{idx:04d}",
                object_type="RowObject",
                text=str(row_dict),
                page_ref_or_sheet_ref=table_ref,
                metadata={"row_index": idx},
            )
        )

    mapping_decisions = []
    for idx, header in enumerate(headers, start=1):
        samples = [row[idx - 1] for row in data_rows[:5] if idx - 1 < len(row)]
        res = resolve_alias(_build_header_bundle(header, samples, modality=modality_l, column_index=idx), pipeline_run_id="ingest", file_fingerprint=_artifact_hash(path))
        mapping_decisions.append(res.decision)

    refs = [obj.object_id for obj in evidence_objects]
    field_claims = [
        _field_claim("site_roster_spreadsheet", modality_l, "pre_field", "site_count", "site_count", len(data_rows), f"site_roster_spreadsheet.{modality_l}.pre", refs),
        _field_claim("site_roster_spreadsheet", modality_l, "pre_field", "site_roster_rows", "site_roster_rows[]", row_claim_value, f"site_roster_spreadsheet.{modality_l}.pre", refs),
    ]
    if any(h.lower() in {"address", "city / state / zip", "city/state/zip"} for h in headers):
        values = []
        for row in data_rows:
            for col, header in enumerate(headers):
                if header.lower() in {"address", "city / state / zip", "city/state/zip"} and col < len(row):
                    values.append(row[col])
        if values:
            field_claims.append(_field_claim("site_roster_spreadsheet", modality_l, "pre_field", "location_details", "location_details[]", values, f"site_roster_spreadsheet.{modality_l}.pre", refs))

    return {
        "role_graph": _role_graph(path, "site_roster_spreadsheet", modality_l, f"{len(data_rows)} roster rows"),
        "evidence_objects": evidence_objects,
        "field_claims": field_claims,
        "mapping_decisions": mapping_decisions,
        "review_flags": [],
    }


__all__ = [
    "ingest_drawing_packet",
    "ingest_site_roster_spreadsheet",
    "ingest_transcript_or_notes",
]
