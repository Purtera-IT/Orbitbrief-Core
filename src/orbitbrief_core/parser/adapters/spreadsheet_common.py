from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping
import warnings

from openpyxl import load_workbook


_NOISE_SHEET_TOKENS = (
    "helper",
    "do not edit",
    "sell rates",
    "selll rates",
    "cost rates",
    "lookup",
)

_RELEVANT_HEADER_TOKENS = (
    "site",
    "location",
    "country",
    "address",
    "region",
    "wave",
    "market",
    "job",
    "description",
    "billing",
    "quantity",
    "qty",
    "duration",
    "term",
    "month",
    "rate type",
    "fe level",
    "service",
    "contact",
    "notes",
    "go live",
    "schedule",
    "project",
)

_RELEVANT_KV_LABEL_TOKENS = (
    "customer",
    "end user",
    "site",
    "location",
    "region",
    "division",
    "project duration",
    "duration",
    "term",
    "billing",
    "service",
    "project type",
    "service category",
    "project summary",
    "deliverable",
    "assumption",
    "responsibil",
    "exclusion",
    "risk",
    "dependen",
    "question",
    "testing",
    "access",
    "material",
)

_FINANCE_HEADER_TOKENS = (
    "revenue",
    "cost",
    "margin",
    "gm",
    "bac",
    "fee",
    "discount",
    "unit sell rate",
    "unit cost rate",
)

_IGNORED_KV_LABELS = (
    "oppty",
    "sales rep",
    "quoted w partner",
    "quoted with partner",
    "channel direct",
    "enterprise technical",
)

_PACKET_TO_CUES: dict[str, tuple[str, ...]] = {
    "scope_packet": ("scope_included",),
    "deliverable_packet": ("deliverable",),
    "site_packet": ("site_location",),
    "quantity_packet": ("quantity",),
    "schedule_packet": ("schedule",),
    "responsibility_packet": ("customer_responsibility",),
}

_FACT_LABEL_BY_FAMILY: dict[str, str] = {
    "scope_packet": "Scope",
    "deliverable_packet": "Deliverable",
    "site_packet": "Site",
    "quantity_packet": "Quantity",
    "schedule_packet": "Schedule",
    "responsibility_packet": "Responsibility",
}


@dataclass(frozen=True, slots=True)
class SpreadsheetBlock:
    sheet_name: str
    row_index: int
    kind: str
    text: str
    metadata: Mapping[str, Any]


@dataclass(frozen=True, slots=True)
class SpreadsheetPreview:
    preview_text: str
    full_text: str
    block_count: int
    relevant_sheet_names: tuple[str, ...]
    skipped_sheet_names: tuple[str, ...]


def normalize_header(text: str) -> str:
    return " ".join(str(text or "").strip().lower().replace("/", " ").replace("_", " ").split())


def coerce_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        text = f"{value:.6f}".rstrip("0").rstrip(".")
        return text
    return str(value).strip()


def is_noise_sheet_name(name: str) -> bool:
    lower = normalize_header(name)
    return any(token in lower for token in _NOISE_SHEET_TOKENS)


def _read_rows(path: Path) -> tuple[tuple[str, list[list[str]]], ...]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
            reader = csv.reader(handle)
            rows = [[coerce_cell_text(cell) for cell in row] for row in reader]
        return (("CSV", rows),)

    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Data Validation extension is not supported")
        workbook = load_workbook(path, data_only=True, read_only=True)
    sheets: list[tuple[str, list[list[str]]]] = []
    for ws in workbook.worksheets:
        rows = [[coerce_cell_text(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
        sheets.append((ws.title, rows))
    return tuple(sheets)


def _non_empty_cells(row: list[str]) -> list[tuple[int, str]]:
    return [(idx, cell.strip()) for idx, cell in enumerate(row) if str(cell).strip()]


def _looks_like_label(text: str) -> bool:
    clean = text.strip()
    if not clean or len(clean) > 64:
        return False
    if clean.isdigit():
        return False
    digit_ratio = sum(ch.isdigit() for ch in clean) / max(1, len(clean))
    if digit_ratio > 0.35:
        return False
    if clean.endswith(":"):
        return True
    return True


def _extract_label_value_pairs(row: list[str]) -> tuple[tuple[str, str, int, int], ...]:
    cells = _non_empty_cells(row)
    pairs: list[tuple[str, str, int, int]] = []
    cursor = 0
    while cursor + 1 < len(cells):
        label_idx, label = cells[cursor]
        value_idx, value = cells[cursor + 1]
        if value_idx - label_idx <= 2 and _looks_like_label(label):
            pairs.append((label.rstrip(":"), value, label_idx, value_idx))
            cursor += 2
            continue
        cursor += 1
    return tuple(pairs)


def _looks_like_header_row(row: list[str]) -> bool:
    cells = _non_empty_cells(row)
    if len(cells) < 4:
        return False
    pairs = _extract_label_value_pairs(row)
    if pairs and any(_looks_like_summary_value(value) for _, value, _, _ in pairs):
        return False
    short_text_cells = 0
    for _, cell in cells:
        digit_ratio = sum(ch.isdigit() for ch in cell) / max(1, len(cell))
        if digit_ratio < 0.25 and len(cell) <= 48:
            short_text_cells += 1
    return short_text_cells >= max(3, int(len(cells) * 0.6))


def _looks_like_summary_value(text: str) -> bool:
    clean = str(text or "").strip()
    if not clean:
        return False
    lower = normalize_header(clean)
    if lower in {"tbd", "yes", "no", "direct", "channel", "partner"}:
        return True
    if re_fullmatch_date(clean):
        return True
    try:
        float(clean.replace(",", ""))
        return True
    except Exception:
        return False


def re_fullmatch_date(text: str) -> bool:
    try:
        datetime.fromisoformat(str(text).strip())
        return True
    except Exception:
        return False


def _kv_label_is_relevant(label: str) -> bool:
    lower = normalize_header(label)
    if not lower:
        return False
    if any(token in lower for token in _IGNORED_KV_LABELS):
        return False
    if any(token in lower for token in _FINANCE_HEADER_TOKENS):
        return False
    return any(token in lower for token in _RELEVANT_KV_LABEL_TOKENS)


def _select_relevant_headers(headers: list[str]) -> tuple[int, ...]:
    selected: list[int] = []
    for idx, header in enumerate(headers):
        lower = normalize_header(header)
        if not lower:
            continue
        if any(token in lower for token in _RELEVANT_HEADER_TOKENS):
            selected.append(idx)
            continue
        if any(token in lower for token in _FINANCE_HEADER_TOKENS):
            continue
        if idx < 4 and len(lower) <= 24:
            selected.append(idx)
    return tuple(dict.fromkeys(selected))


def _clean_scope_text(text: str) -> str:
    cleaned = " ".join(text.replace("_", " ").split()).strip()
    cleaned = cleaned.removesuffix(" - LABOR").strip()
    return cleaned


def _compact_quantity(value: str, unit_hint: str | None = None) -> str:
    base = str(value).strip()
    if not base:
        return ""
    if unit_hint:
        lowered = unit_hint.lower()
        if "month" in lowered:
            return f"{base} months"
        if "week" in lowered:
            return f"{base} weeks"
        if "day" in lowered:
            return f"{base} days"
        if lowered in {"ea", "each", "unit", "units"}:
            return f"{base} units"
        if unit_hint.strip():
            return f"{base} {unit_hint.strip()}".strip()
    return base


def _claim_overrides_for_kv(label: str, value: str) -> dict[str, str]:
    lower = normalize_header(label)
    overrides: dict[str, str] = {}
    if not value:
        return overrides
    if any(token in lower for token in ("qty of sites", "site count", "number of sites", "qnty of sites")):
        compact = _compact_quantity(value, "sites")
        overrides["site_packet"] = compact
        overrides["quantity_packet"] = compact
    elif any(token in lower for token in ("project duration", "term", "duration")):
        unit = "months" if "month" in lower else None
        compact = _compact_quantity(value, unit)
        overrides["schedule_packet"] = compact
        overrides["quantity_packet"] = compact
    elif "billing type" in lower:
        overrides["schedule_packet"] = value
    elif "division" in lower or "service category" in lower or "project type" in lower:
        overrides["scope_packet"] = value
    return overrides


def _claim_hints_for_kv(label: str, value: str) -> tuple[str, ...]:
    lower = normalize_header(label)
    hints: list[str] = []
    if not value:
        return ()
    if any(token in lower for token in ("customer", "end user")):
        hints.append("customer_identity")
    if "project summary" in lower:
        hints.append("project_summary")
    if any(token in lower for token in ("billing type", "billing")):
        hints.append("commercial_structure_claim")
    if any(token in lower for token in ("qty of sites", "site count", "number of sites", "total sites")):
        hints.append("site_count_claim")
    if any(token in lower for token in ("contact", "email", "phone", "title")) and ("@" in value or len(value.split()) >= 2):
        hints.append("contact_claim")
    return tuple(dict.fromkeys(hints))


def _claim_overrides_for_row(row_map: Mapping[str, str]) -> dict[str, str]:
    lowered = {normalize_header(key): value for key, value in row_map.items() if str(key).strip() and str(value).strip()}
    overrides: dict[str, str] = {}
    site = lowered.get("site") or lowered.get("site name") or lowered.get("location")
    job = lowered.get("job description") or lowered.get("scope") or lowered.get("description")
    billing = lowered.get("billing type") or lowered.get("billing")
    quantity = (
        lowered.get("unit sell quantity")
        or lowered.get("unit cost quantity")
        or lowered.get("quantity")
        or lowered.get("qty")
        or lowered.get("est. units")
    )
    unit_hint = lowered.get("labor rate type (if applicable)") or lowered.get("labor rate type") or lowered.get("rate type")
    if site:
        overrides["site_packet"] = site
    if job:
        cleaned_job = _clean_scope_text(job)
        if cleaned_job:
            overrides["scope_packet"] = cleaned_job
    if quantity:
        compact = _compact_quantity(quantity, unit_hint)
        if compact:
            overrides["quantity_packet"] = compact
    if billing:
        schedule_bits = [billing]
        compact = overrides.get("quantity_packet")
        if compact and any(token in compact.lower() for token in ("month", "week", "day")):
            schedule_bits.append(compact)
        overrides["schedule_packet"] = "; ".join(dict.fromkeys(bit for bit in schedule_bits if bit))
    return overrides


def _claim_hints_for_row(row_map: Mapping[str, str]) -> tuple[str, ...]:
    lowered = {normalize_header(key): value for key, value in row_map.items() if str(key).strip() and str(value).strip()}
    hints: list[str] = []
    if any(key in lowered for key in ("customer", "end user")):
        hints.append("customer_identity")
    if any(key in lowered for key in ("job description", "scope", "description")):
        hints.append("project_summary")
    if any(key in lowered for key in ("billing type", "billing")):
        hints.append("commercial_structure_claim")
    if any(key in lowered for key in ("quantity", "qty", "unit sell quantity", "est. units")) and any(key in lowered for key in ("site", "site name", "location")):
        hints.append("site_count_claim")
    if any(key in lowered for key in ("contact", "email", "phone", "title", "name")):
        hints.append("contact_claim")
    return tuple(dict.fromkeys(hints))


def _metadata_for_overrides(overrides: Mapping[str, str], claim_hints: tuple[str, ...] = ()) -> dict[str, Any]:
    packet_families = sorted(overrides)
    parser_cues: list[str] = []
    for family in packet_families:
        parser_cues.extend(_PACKET_TO_CUES.get(family, ()))
    metadata = {
        "claim_body_overrides": dict(overrides),
        "parser_cues": sorted(set(parser_cues)),
        "packet_families": packet_families,
    }
    if claim_hints:
        metadata["target_claim_family_hints"] = list(dict.fromkeys(claim_hints))
    return metadata


def _row_fact_blocks(sheet_name: str, row_index: int, row_map: Mapping[str, str], overrides: Mapping[str, str], claim_hints: tuple[str, ...] = ()) -> Iterator[SpreadsheetBlock]:
    for family, value in overrides.items():
        clean_value = str(value).strip()
        if not clean_value:
            continue
        label = _FACT_LABEL_BY_FAMILY.get(family, family.replace("_packet", "").replace("_", " ").title())
        metadata = {
            "kind": "spreadsheet_fact",
            "sheet_name": sheet_name,
            "row_index": row_index,
            "row_values": dict(row_map),
            "fact_family": family,
        }
        metadata.update(_metadata_for_overrides({family: clean_value}, claim_hints=claim_hints))
        yield SpreadsheetBlock(
            sheet_name=sheet_name,
            row_index=row_index,
            kind="spreadsheet_fact",
            text=f"{label}: {clean_value}",
            metadata=metadata,
        )


def _iter_relevant_blocks(sheet_name: str, rows: list[list[str]]) -> Iterator[SpreadsheetBlock]:
    if not rows:
        return
    header_row_index: int | None = None
    headers: list[str] = []
    relevant_table_columns: tuple[int, ...] = ()
    for idx, row in enumerate(rows, start=1):
        if header_row_index is None and _looks_like_header_row(row):
            header_row_index = idx
            headers = [cell.strip() for cell in row]
            relevant_table_columns = _select_relevant_headers(headers)
            continue

        pairs = _extract_label_value_pairs(row)
        if pairs and (header_row_index is None or idx < header_row_index):
            for label, value, label_col, value_col in pairs:
                if not _kv_label_is_relevant(label):
                    continue
                overrides = _claim_overrides_for_kv(label, value)
                claim_hints = _claim_hints_for_kv(label, value)
                if not overrides and not claim_hints:
                    continue
                metadata = {
                    "kind": "spreadsheet_kv",
                    "sheet_name": sheet_name,
                    "row_index": idx,
                    "label": label,
                    "value": value,
                    "label_column": label_col + 1,
                    "value_column": value_col + 1,
                    "normalized_label": normalize_header(label),
                }
                metadata.update(_metadata_for_overrides(overrides, claim_hints=claim_hints))
                yield SpreadsheetBlock(
                    sheet_name=sheet_name,
                    row_index=idx,
                    kind="spreadsheet_kv",
                    text=f"{label}: {value}",
                    metadata=metadata,
                )
            continue

        if header_row_index is None or idx <= header_row_index:
            continue

        non_empty = _non_empty_cells(row)
        if len(non_empty) < 2:
            continue

        selected_indices = relevant_table_columns or tuple(i for i, _ in non_empty[:8])
        row_map = {
            headers[col]: row[col].strip()
            for col in selected_indices
            if col < len(headers) and col < len(row) and headers[col].strip() and row[col].strip()
        }
        if not row_map:
            continue
        overrides = _claim_overrides_for_row(row_map)
        claim_hints = _claim_hints_for_row(row_map)
        parts = [f"{key}: {value}" for key, value in row_map.items()]
        metadata = {
            "kind": "spreadsheet_row",
            "sheet_name": sheet_name,
            "row_index": idx,
            "row_values": dict(row_map),
            "headers": [headers[col] for col in selected_indices if col < len(headers)],
        }
        metadata.update(_metadata_for_overrides(overrides, claim_hints=claim_hints))
        yield SpreadsheetBlock(
            sheet_name=sheet_name,
            row_index=idx,
            kind="spreadsheet_row",
            text="; ".join(parts),
            metadata=metadata,
        )
        yield from _row_fact_blocks(sheet_name, idx, row_map, overrides, claim_hints=claim_hints)


def extract_spreadsheet_blocks(path: Path, *, max_sheets: int = 4, max_blocks_per_sheet: int = 24) -> tuple[SpreadsheetBlock, ...]:
    blocks: list[SpreadsheetBlock] = []
    relevant_sheets = 0
    for sheet_name, rows in _read_rows(path):
        if is_noise_sheet_name(sheet_name):
            continue
        relevant_sheets += 1
        sheet_blocks = list(_iter_relevant_blocks(sheet_name, rows))
        blocks.extend(sheet_blocks[:max_blocks_per_sheet])
        if relevant_sheets >= max_sheets:
            break
    return tuple(blocks)


def build_spreadsheet_preview(path: Path, *, max_preview_chars: int = 12000) -> SpreadsheetPreview:
    preview_lines: list[str] = []
    full_lines: list[str] = []
    relevant_sheet_names: list[str] = []
    skipped_sheet_names: list[str] = []
    for sheet_name, rows in _read_rows(path):
        if is_noise_sheet_name(sheet_name):
            skipped_sheet_names.append(sheet_name)
            continue
        relevant_sheet_names.append(sheet_name)
        full_lines.append(f"Sheet: {sheet_name}")
        preview_lines.append(f"Sheet: {sheet_name}")
        blocks = list(_iter_relevant_blocks(sheet_name, rows))
        for block in blocks:
            line = f"{block.text}"
            full_lines.append(line)
            if len("\n".join(preview_lines)) < max_preview_chars:
                preview_lines.append(line)
    full_text = "\n".join(full_lines).strip()
    preview_text = "\n".join(preview_lines).strip()[:max_preview_chars]
    return SpreadsheetPreview(
        preview_text=preview_text,
        full_text=full_text,
        block_count=len(extract_spreadsheet_blocks(path)),
        relevant_sheet_names=tuple(relevant_sheet_names),
        skipped_sheet_names=tuple(skipped_sheet_names),
    )
