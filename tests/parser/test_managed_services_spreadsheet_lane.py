from __future__ import annotations

from dataclasses import dataclass

from openpyxl import Workbook

from orbitbrief_core.parser.router import RouterInput
from orbitbrief_core.parser.runtime import parse_extract_and_postprocess
from orbitbrief_core.runtime_spine.extractors.packet_to_claims import PacketExtractionContext, extract_claims_from_packet


@dataclass(frozen=True, slots=True)
class _ManifestStub:
    pack_id: str = "professional_services_text"
    role_id: str = "transcript_or_notes"


@dataclass(frozen=True, slots=True)
class _CompiledPackStub:
    manifest: _ManifestStub
    parser_profiles: dict


def _compiled_pack_stub() -> _CompiledPackStub:
    rows = [
        {"modality": "txt", "parser_profile_id": "parser:professional_services_text:txt"},
        {"modality": "md", "parser_profile_id": "parser:professional_services_text:md"},
        {"modality": "docx", "parser_profile_id": "parser:professional_services_text:docx"},
        {"modality": "email_export", "parser_profile_id": "parser:professional_services_text:email_export"},
        {"modality": "pdf_text", "parser_profile_id": "parser:professional_services_text:pdf_text"},
        {"modality": "pdf_ocr", "parser_profile_id": "parser:professional_services_text:pdf_ocr"},
        {"modality": "xlsx", "parser_profile_id": "parser:professional_services_text:xlsx"},
        {"modality": "csv", "parser_profile_id": "parser:professional_services_text:csv"},
    ]
    return _CompiledPackStub(manifest=_ManifestStub(), parser_profiles={"rows": rows})


def test_template_schema_prompt_package_is_routed_to_intake_only() -> None:
    compiled_pack = _compiled_pack_stub()
    prompt_package_text = (
        'Prompt Package JSON\n'
        '{"system_prompt": "Use PurTera exactly as written.", '
        '"user_prompt_template": "Generate a JSON object with {{project_type}}, {{site_count}}, {{location_details}}, and {{deliverables_needed_json}}. '
        'Return valid JSON only.", '
        '"output_schema": {"type": "object", "required": ["scope_overview"], '
        '"additionalProperties": false}}'
    )
    router_input = RouterInput(
        doc_id="prompt_package_docx_001",
        filename="prompt_package.docx",
        raw_text_preview=prompt_package_text,
        metadata={"raw_text": prompt_package_text},
    )
    result = parse_extract_and_postprocess(router_input=router_input, compiled_pack=compiled_pack)

    assert result.pipeline_state == "intake_only"
    assert "template_schema_artifact" in result.reason_codes
    assert result.emits_business_claims is False
    assert result.postprocess_result["summary"]["claims_emitted_count"] == 0


def test_managed_services_spreadsheet_lane_recovers_site_scope_quantity_and_schedule(tmp_path) -> None:
    workbook_path = tmp_path / "deal_kit.xlsx"
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Deal Kit"
    summary.append(["Deal Summary", None, None, "Overall Deal Kit Summary", None])
    summary.append(["Customer", "Musick, Peeler & Garrett", None, None, None])
    summary.append(["QTY of Sites", 5, None, None, None])
    summary.append(["Division", "EUC / Helpdesk Support", None, None, None])
    summary.append(["Project Duration (Months)", 12, None, None, None])
    summary.append(["Billing Type", "Fixed Fee - Monthly Billing", None, None, None])

    gantt = workbook.create_sheet("Gantt Financials")
    gantt.append(
        [
            "Site",
            "Country",
            "FE Level",
            "Job Description",
            "Billing Type",
            "Country Multiplier",
            "Unit Sell Quantity",
            "Labor Rate Type\n(If Applicable)",
        ]
    )
    gantt.append(
        [
            "Los Angeles HQ",
            "USA",
            "L2 EUC",
            "Dedicated Full-Time Onsite/Remote IT Support Resource - LABOR",
            "Fixed Fee - Monthly Billing",
            1,
            12,
            "Per Month",
        ]
    )

    helper = workbook.create_sheet("Helper - Do not Edit")
    helper.append([None, "Lookup Row"])
    workbook.save(workbook_path)

    compiled_pack = _compiled_pack_stub()
    router_input = RouterInput(
        doc_id="deal_kit_001",
        filename=str(workbook_path),
        raw_text_preview="",
        metadata={"path": str(workbook_path), "raw_text": ""},
    )
    result = parse_extract_and_postprocess(router_input=router_input, compiled_pack=compiled_pack)

    assert result.pipeline_state == "extract"
    assert result.parse_runtime_result.parse_plan.metadata["modality"] == "xlsx"
    values = [
        str(claim["candidate_value"])
        for claim in result.postprocess_result["normalized_output"]["field_claims"]
    ]
    assert any("Los Angeles HQ" in value for value in values)
    assert any("Dedicated Full-Time Onsite/Remote IT Support Resource" in value for value in values)
    assert any("5 sites" in value for value in values)
    assert any("12 months" in value for value in values)
    review_messages = [flag.message for flag in result.parse_runtime_result.document_parse.review_flags]
    assert any("helper/lookup sheets were ignored" in message.lower() for message in review_messages)


def test_packet_to_claims_prefers_bullets_over_section_headings() -> None:
    packet = {
        "packet_id": "packet:deliverable_heading:0001",
        "span_ids": ("span_heading", "span_bullet"),
        "primary_span_id": "span_heading",
        "confidence": 0.79,
        "evidence_rows": [
            {
                "span_id": "span_heading",
                "text": "DELIVERABLES",
                "normalized_text": "deliverables",
                "section_path": ["ROOT", "DELIVERABLES"],
                "parser_cues": ["deliverable"],
                "packet_families": ["deliverable_packet"],
                "authority_score": 0.9,
                "metadata": {"kind": "heading"},
            },
            {
                "span_id": "span_bullet",
                "text": "Provision of one assigned full-time support resource for the duration of the engagement.",
                "normalized_text": "provision of one assigned full-time support resource for the duration of the engagement",
                "section_path": ["ROOT", "DELIVERABLES"],
                "parser_cues": ["deliverable"],
                "packet_families": ["deliverable_packet"],
                "authority_score": 0.9,
                "metadata": {"kind": "bullet"},
            },
        ],
        "metadata": {
            "packet_family": "deliverable_packet",
            "uncertainty_markers": [],
            "packet_diagnostic": {"included": [{"span_id": "span_heading"}, {"span_id": "span_bullet"}]},
        },
    }
    claims, diagnostics = extract_claims_from_packet(
        packet,
        PacketExtractionContext(role_id="transcript_or_notes", modality="docx"),
    )

    assert not diagnostics
    assert len(claims) == 1
    assert claims[0].claim_body == "Provision of one assigned full-time support resource for the duration of the engagement"
    assert claims[0].metadata["semantic_source_span_id"] == "span_bullet"


def test_low_signal_legal_sections_are_suppressed() -> None:
    packet = {
        "packet_id": "packet:acceptance:0001",
        "span_ids": ("span_legal",),
        "primary_span_id": "span_legal",
        "confidence": 0.78,
        "evidence_rows": [
            {
                "span_id": "span_legal",
                "text": "By Customer signature below, Customer accepts this SOW and agrees to the terms and conditions of the Agreement.",
                "normalized_text": "by customer signature below customer accepts this sow and agrees to the terms and conditions of the agreement",
                "section_path": ["ROOT", "ACCEPTANCE CRITERIA"],
                "parser_cues": ["assumption", "customer_responsibility"],
                "packet_families": ["scope_packet"],
                "authority_score": 0.82,
                "metadata": {"kind": "paragraph"},
            },
        ],
        "metadata": {
            "packet_family": "scope_packet",
            "uncertainty_markers": [],
            "packet_diagnostic": {"included": [{"span_id": "span_legal"}]},
        },
    }
    claims, diagnostics = extract_claims_from_packet(
        packet,
        PacketExtractionContext(role_id="transcript_or_notes", modality="docx"),
    )

    assert claims == ()
    assert any(item.code in {"low_signal_section_suppressed", "semantic_signal_too_weak"} for item in diagnostics)


def test_parser_architecture_pdf_is_routed_to_intake_only() -> None:
    compiled_pack = _compiled_pack_stub()
    architecture_text = (
        "OrbitBrief parser architecture\n"
        "multimodal evidence compiler\n"
        "graph-backed packet neighborhoods\n"
        "bounded narrative extraction\n"
        "projection + deterministic postprocess\n"
        "compiled runtime policy\n"
        "Qwen embeddings and packetizer\n"
        "Layer 1 Layer 2 Layer 3 Layer 4 Layer 5 Layer 6 Layer 7 Layer 8 Layer 9 Layer 10 Layer 11"
    )
    router_input = RouterInput(
        doc_id="parser_architecture_pdf_001",
        filename="ParserOs.pdf",
        raw_text_preview=architecture_text,
        metadata={"raw_text": architecture_text},
    )
    result = parse_extract_and_postprocess(router_input=router_input, compiled_pack=compiled_pack)

    assert result.pipeline_state == "intake_only"
    assert "meta_reference_artifact" in result.reason_codes
    assert result.postprocess_result["summary"]["claims_emitted_count"] == 0


def test_customer_identity_semantic_lift_rejects_customer_hyphenated_phrases() -> None:
    packet = {
        "packet_id": "packet:customer_identity:0001",
        "span_ids": ("span_bad", "span_good"),
        "primary_span_id": "span_bad",
        "confidence": 0.78,
        "evidence_rows": [
            {
                "span_id": "span_bad",
                "text": "Current platform details, ticketing tooling, access methods, and any customer-specific application requirements shall be confirmed during onboarding and transition.",
                "normalized_text": "current platform details ticketing tooling access methods and any customer specific application requirements shall be confirmed during onboarding and transition",
                "section_path": ["ROOT", "ASSUMPTIONS"],
                "packet_families": ["schedule_packet"],
                "authority_score": 0.84,
                "metadata": {"kind": "paragraph"},
            },
            {
                "span_id": "span_good",
                "text": "PurTera shall provide a dedicated full-time technical support resource for Musick, Peeler & Garrett to perform onsite services primarily from the Los Angeles office.",
                "normalized_text": "purtera shall provide a dedicated full time technical support resource for musick peeler and garrett to perform onsite services primarily from the los angeles office",
                "section_path": ["ROOT", "PROJECT OVERVIEW"],
                "packet_families": ["scope_packet"],
                "authority_score": 0.91,
                "metadata": {"kind": "paragraph"},
            },
        ],
        "metadata": {
            "packet_family": "schedule_packet",
            "uncertainty_markers": ["family_conflict"],
            "packet_diagnostic": {"anchor": {"family_hints": ["scope_packet"]}},
        },
    }
    claims, diagnostics = extract_claims_from_packet(
        packet,
        PacketExtractionContext(role_id="transcript_or_notes", modality="docx"),
    )

    customer_claims = [claim for claim in claims if claim.claim_family == "customer_identity"]
    assert customer_claims
    assert all(claim.claim_body == "Musick, Peeler & Garrett" for claim in customer_claims)
    assert all("customer-specific" not in claim.claim_body.lower() for claim in customer_claims)


def test_site_packet_suppresses_internal_anchor_fallback_strings() -> None:
    packet = {
        "packet_id": "packet:site_anchor_fallback:0001",
        "span_ids": ("span_site",),
        "primary_span_id": "span_site",
        "confidence": 0.82,
        "evidence_rows": [
            {
                "span_id": "span_site",
                "text": "Site: anchor=span:doc:0001 supports=5",
                "normalized_text": "site: anchor=span:doc:0001 supports=5",
                "section_path": ["ROOT", "PROJECT OVERVIEW"],
                "parser_cues": ["site_location"],
                "packet_families": ["site_packet"],
                "authority_score": 0.91,
                "metadata": {"kind": "paragraph"},
            },
        ],
        "metadata": {
            "packet_family": "site_packet",
            "uncertainty_markers": [],
            "packet_diagnostic": {"included": [{"span_id": "span_site"}]},
        },
    }
    claims, diagnostics = extract_claims_from_packet(
        packet,
        PacketExtractionContext(role_id="transcript_or_notes", modality="docx"),
    )

    site_claims = [claim for claim in claims if claim.claim_family == "site_location_claim"]
    assert site_claims == []
    assert any(item.code in {"site_location_not_specific", "semantic_signal_too_weak"} for item in diagnostics)


def test_customer_contact_tables_emit_only_customer_contact_rows() -> None:
    packet = {
        "packet_id": "packet:contact_rows:0001",
        "span_ids": ("span_vendor_header", "span_vendor", "span_customer_header", "span_customer", "span_signature"),
        "primary_span_id": "span_customer",
        "confidence": 0.81,
        "evidence_rows": [
            {
                "span_id": "span_vendor_header",
                "text": "FULL NAME | JOB TITLE | EMAIL ADDRESS",
                "normalized_text": "full name | job title | email address",
                "section_path": ["ROOT", "PURTERA SALES CONTACTS"],
                "parser_cues": [],
                "packet_families": ["responsibility_packet"],
                "authority_score": 0.82,
                "metadata": {"kind": "table_row", "contact_scope": "vendor", "target_claim_family_hints": ["contact_claim"]},
            },
            {
                "span_id": "span_vendor",
                "text": "Chase Smith | Director of Operations | chase@purtera-it.com",
                "normalized_text": "chase smith | director of operations | chase@purtera-it.com",
                "section_path": ["ROOT", "PURTERA SALES CONTACTS"],
                "parser_cues": [],
                "packet_families": ["responsibility_packet"],
                "authority_score": 0.9,
                "metadata": {
                    "kind": "table_row",
                    "contact_scope": "vendor",
                    "target_claim_family_hints": ["contact_claim"],
                    "row_values": {
                        "FULL NAME": "Chase Smith",
                        "JOB TITLE": "Director of Operations",
                        "EMAIL ADDRESS": "chase@purtera-it.com",
                    },
                },
            },
            {
                "span_id": "span_customer_header",
                "text": "FULL NAME | JOB TITLE | EMAIL ADDRESS",
                "normalized_text": "full name | job title | email address",
                "section_path": ["ROOT", "CUSTOMER CONTACTS"],
                "parser_cues": [],
                "packet_families": ["responsibility_packet"],
                "authority_score": 0.82,
                "metadata": {"kind": "table_row", "contact_scope": "customer", "target_claim_family_hints": ["contact_claim"]},
            },
            {
                "span_id": "span_customer",
                "text": "Jamie Barrios | Director of Operations",
                "normalized_text": "jamie barrios | director of operations",
                "section_path": ["ROOT", "CUSTOMER CONTACTS"],
                "parser_cues": [],
                "packet_families": ["responsibility_packet"],
                "authority_score": 0.9,
                "metadata": {
                    "kind": "table_row",
                    "contact_scope": "customer",
                    "target_claim_family_hints": ["contact_claim"],
                    "row_values": {
                        "FULL NAME": "Jamie Barrios",
                        "JOB TITLE": "Director of Operations",
                    },
                },
            },
            {
                "span_id": "span_signature",
                "text": "Name:\nDate: | Name:\nDate:",
                "normalized_text": "name: date: | name: date:",
                "section_path": ["ROOT", "ACCEPTANCE CRITERIA"],
                "parser_cues": [],
                "packet_families": ["responsibility_packet"],
                "authority_score": 0.7,
                "metadata": {"kind": "table_row"},
            },
        ],
        "metadata": {
            "packet_family": "responsibility_packet",
            "uncertainty_markers": ["family_conflict"],
            "packet_diagnostic": {"included": [{"span_id": "span_customer"}]},
        },
    }
    claims, diagnostics = extract_claims_from_packet(
        packet,
        PacketExtractionContext(role_id="transcript_or_notes", modality="docx"),
    )

    contact_claims = [claim for claim in claims if claim.claim_family == "contact_claim"]
    assert [claim.claim_body for claim in contact_claims] == ["Jamie Barrios - Director of Operations"]
    assert all("Chase Smith" not in claim.claim_body for claim in contact_claims)
    assert all("Name:" not in claim.claim_body for claim in contact_claims)
    assert any(item.code in {"direct_claim_extracted", "non_customer_contact_suppressed", "contact_header_suppressed"} for item in diagnostics)
